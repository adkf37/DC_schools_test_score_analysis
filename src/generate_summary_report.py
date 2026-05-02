"""
Formatted Excel Summary Report — DC Schools Test Score Analysis.

Reads from existing output CSVs and produces a formatted multi-sheet Excel
workbook that summarises key findings for policy stakeholders.

Sheets produced:
  1. Executive Summary    — Citywide headline statistics
  2. Top Growth (ELA)     — Top schools by avg cohort growth, ELA, All Students
  3. Top Growth (Math)    — Top schools by avg cohort growth, Math, All Students
  4. Top Equity Schools   — Schools most effective at narrowing equity gaps
  5. Proficiency Trends   — Citywide avg proficiency by year and subject
  6. School Directory     — All schools with combined growth + equity metrics
  7. Geographic Equity    — Performance by DC quadrant (NE / NW / SE / SW)
  8. YoY Growth           — Same-grade year-over-year growth by subject & transition
  9. COVID Recovery       — 2019→2022 COVID impact and 2022→2024 recovery per school
 10. School Trajectories  — Long-run linear trend slope and class per school (2016–2024)
 11. School Types         — Avg proficiency, COVID recovery, and cohort growth by school type
 12. Grade Levels         — Avg proficiency, COVID recovery, and YoY growth by grade (3–8, HS)
 13. Subgroups            — Avg proficiency, COVID recovery, and YoY growth by student subgroup
 14. Consistency          — Year-to-year performance stability (CV, range) by school
 15. Performance Index    — Multi-metric composite score (proficiency, growth, recovery, trajectory)

Output:
    output_data/summary_report.xlsx

Usage:
    python src/generate_summary_report.py

Dependencies:
    - output_data/cohort_growth_summary.csv  (produced by analyze_cohort_growth.py)
    - output_data/school_rankings.csv        (produced by generate_school_rankings.py)
    - output_data/school_equity_rankings.csv (produced by generate_school_rankings.py)
    - output_data/equity_gap_summary.csv     (produced by equity_gap_analysis.py)
    - output_data/proficiency_trends.csv     (produced by proficiency_trend_analysis.py)
    - output_data/geographic_equity_by_quadrant.csv (produced by geographic_equity_analysis.py)
    - output_data/yoy_growth_summary.csv     (produced by yoy_growth_analysis.py, optional)
    - output_data/covid_recovery_summary.csv (produced by covid_recovery_analysis.py, optional)
    - output_data/school_trajectory_classification.csv (produced by school_trajectory_analysis.py, optional)
    - output_data/school_type_summary.csv    (produced by school_type_analysis.py, optional)
    - output_data/grade_level_summary.csv    (produced by grade_level_analysis.py, optional)
    - output_data/subgroup_summary.csv       (produced by subgroup_trend_analysis.py, optional)
    - output_data/school_consistency.csv     (produced by school_consistency_analysis.py, optional)
    - output_data/school_performance_index.csv (produced by school_performance_index.py, optional)
"""
import os
import sys
import pandas as pd

# ── Paths ─────────────────────────────────────────────────────────────────────
CURRENT_PATH = os.path.abspath(os.path.dirname(__file__))
OUTPUT_PATH = os.path.abspath(os.path.join(CURRENT_PATH, '..', 'output_data'))

COHORT_SUMMARY_FILE = os.path.join(OUTPUT_PATH, 'cohort_growth_summary.csv')
RANKINGS_FILE = os.path.join(OUTPUT_PATH, 'school_rankings.csv')
EQUITY_RANKINGS_FILE = os.path.join(OUTPUT_PATH, 'school_equity_rankings.csv')
EQUITY_SUMMARY_FILE = os.path.join(OUTPUT_PATH, 'equity_gap_summary.csv')
TRENDS_FILE = os.path.join(OUTPUT_PATH, 'proficiency_trends.csv')
GEO_EQUITY_FILE = os.path.join(OUTPUT_PATH, 'geographic_equity_by_quadrant.csv')
YOY_SUMMARY_FILE = os.path.join(OUTPUT_PATH, 'yoy_growth_summary.csv')
COVID_RECOVERY_FILE = os.path.join(OUTPUT_PATH, 'covid_recovery_summary.csv')
TRAJECTORY_FILE = os.path.join(OUTPUT_PATH, 'school_trajectory_classification.csv')
SCHOOL_TYPE_FILE = os.path.join(OUTPUT_PATH, 'school_type_summary.csv')
GRADE_LEVEL_FILE = os.path.join(OUTPUT_PATH, 'grade_level_summary.csv')
SUBGROUP_SUMMARY_FILE = os.path.join(OUTPUT_PATH, 'subgroup_summary.csv')
CONSISTENCY_FILE = os.path.join(OUTPUT_PATH, 'school_consistency.csv')
PERFORMANCE_INDEX_FILE = os.path.join(OUTPUT_PATH, 'school_performance_index.csv')
REPORT_FILE = os.path.join(OUTPUT_PATH, 'summary_report.xlsx')

# Top-N schools per subject shown in sheet 2 & 3
TOP_N = 25

# Colour palette
HEADER_BG = '1F4E79'      # dark blue
HEADER_FG = 'FFFFFF'      # white text
ALT_ROW_BG = 'D6E4F0'     # light blue alternating row
POSITIVE_FG = '1E6823'    # dark green — positive growth
NEGATIVE_FG = 'A50026'    # dark red — negative growth
TITLE_BG = '2E75B6'       # medium blue — section headers
TITLE_FG = 'FFFFFF'       # white
SUBHEADER_BG = 'BDD7EE'   # pale blue — subheader rows

ALL_STUDENTS_LABELS = {'All Students', 'All', 'Total'}


# ═════════════════════════════════════════════════════════════════════════════
# Data loading helpers
# ═════════════════════════════════════════════════════════════════════════════

def _load(path: str, label: str) -> pd.DataFrame:
    if not os.path.isfile(path):
        print(f'ERROR: {path} not found.\nRun the pipeline scripts first.')
        sys.exit(1)
    df = pd.read_csv(path)
    print(f'  Loaded {label}: {len(df):,} rows')
    return df


# ═════════════════════════════════════════════════════════════════════════════
# Sheet builders
# ═════════════════════════════════════════════════════════════════════════════

def _build_exec_summary(cohort_summary: pd.DataFrame,
                        rankings: pd.DataFrame,
                        equity_rankings: pd.DataFrame,
                        trends: pd.DataFrame) -> pd.DataFrame:
    """
    Compute a set of citywide headline statistics to place on the
    Executive Summary sheet.
    """
    rows = []

    # ── Dataset scope ──────────────────────────────────────────────────────
    rows.append(('Section', 'Dataset Scope', ''))
    n_schools = cohort_summary['School Name'].nunique()
    rows.append(('', 'Schools with cohort data', n_schools))

    if 'latest_baseline_year' in cohort_summary.columns:
        min_yr = int(cohort_summary['latest_baseline_year'].min())
        max_yr = int(cohort_summary['latest_followup_year'].max())
        rows.append(('', 'Year range', f'{min_yr} – {max_yr}'))

    total_transitions = int(cohort_summary['n_transitions'].sum())
    rows.append(('', 'Total cohort transitions analysed', f'{total_transitions:,}'))

    # ── Citywide growth ────────────────────────────────────────────────────
    rows.append(('', '', ''))
    rows.append(('Section', 'Citywide Cohort Growth (All Students)', ''))

    all_students_mask = cohort_summary['Student Group Value'].isin(ALL_STUDENTS_LABELS)
    city_df = cohort_summary[all_students_mask]

    for subj in ('ELA', 'Math'):
        sub = city_df[city_df['Subject'] == subj]
        if sub.empty:
            continue
        avg = sub['avg_pp_growth'].mean()
        pct_pos = (sub['avg_pp_growth'] > 0).mean() * 100
        rows.append(('', f'{subj} — avg cohort growth (pp)', f'{avg:+.1f}'))
        rows.append(('', f'{subj} — schools with positive growth', f'{pct_pos:.0f}%'))

    # ── Statistical significance ───────────────────────────────────────────
    rows.append(('', '', ''))
    rows.append(('Section', 'Statistical Significance', ''))
    if 'pct_significant_transitions' in cohort_summary.columns:
        for subj in ('ELA', 'Math'):
            sub = city_df[city_df['Subject'] == subj]
            if sub.empty:
                continue
            pct_sig = sub['pct_significant_transitions'].mean()
            rows.append(('', f'{subj} — avg % significant transitions (p < 0.05)',
                         f'{pct_sig:.0f}%'))

    # ── Equity gap ─────────────────────────────────────────────────────────
    if not equity_rankings.empty and 'avg_gap_change' in equity_rankings.columns:
        rows.append(('', '', ''))
        rows.append(('Section', 'Equity Gap Summary', ''))
        for subj in ('ELA', 'Math'):
            sub = equity_rankings[equity_rankings['Subject'] == subj]
            if sub.empty:
                continue
            n_narrowing = (sub['avg_gap_change'] > 0).sum()
            pct_narrowing = n_narrowing / len(sub) * 100
            rows.append(('', f'{subj} — schools narrowing equity gaps',
                         f'{n_narrowing} / {len(sub)} ({pct_narrowing:.0f}%)'))

    # ── Proficiency trends ────────────────────────────────────────────────
    if not trends.empty:
        rows.append(('', '', ''))
        rows.append(('Section', 'Citywide Proficiency Trends (All Students)', ''))
        city_trends = trends[trends['School Name'] == 'DC Public Schools']
        if city_trends.empty:
            city_trends = trends[trends['Student Group Value'].isin(ALL_STUDENTS_LABELS)]

        for subj in ('ELA', 'Math'):
            sub = city_trends[city_trends['Subject'] == subj]
            if sub.empty:
                continue
            by_year = (sub.groupby('year')['proficiency_pct']
                         .mean()
                         .sort_index()
                         .dropna())
            if len(by_year) >= 2:
                years = list(by_year.index)
                vals = list(by_year.values)
                rows.append(('', f'{subj} — {int(years[0])} proficiency',
                             f'{vals[0]:.1f}%'))
                rows.append(('', f'{subj} — {int(years[-1])} proficiency',
                             f'{vals[-1]:.1f}%'))
                delta = vals[-1] - vals[0]
                rows.append(('', f'{subj} — change over period',
                             f'{delta:+.1f} pp'))

    # ── Data caveats ──────────────────────────────────────────────────────
    rows.append(('', '', ''))
    rows.append(('Section', 'Data Notes and Caveats', ''))
    rows.append(('', 'Source', 'DC OSSE wide-format workbooks 2015-16 – 2023-24'))
    rows.append(('', 'Minimum cohort size', 'N >= 10 students per transition'))
    rows.append(('', 'Significance test', 'Two-proportion z-test, alpha = 0.05 (no multiple-comparison correction)'))
    rows.append(('', 'School-year gap', 'No 2019–2022 transitions (COVID-19 disruption)'))
    rows.append(('', 'Coverage caveat', '2024-25 data not yet in repo; charter vs. DCPS split not available'))

    return pd.DataFrame(rows, columns=['Type', 'Metric', 'Value'])


def _build_top_growth(rankings: pd.DataFrame, subject: str, n: int) -> pd.DataFrame:
    """Top-N schools by avg cohort growth for a given subject."""
    sub = rankings[rankings['Subject'] == subject].copy()
    sub = sub.sort_values('avg_pp_growth', ascending=False).head(n)
    sub = sub.reset_index(drop=True)

    out_cols = ['rank', 'School Name', 'avg_pp_growth', 'n_transitions']
    if 'pct_significant_transitions' in sub.columns:
        out_cols.append('pct_significant_transitions')

    out = sub[[c for c in out_cols if c in sub.columns]].copy()

    rename = {
        'rank': 'Rank',
        'School Name': 'School Name',
        'avg_pp_growth': 'Avg Growth (pp)',
        'n_transitions': 'No. Transitions',
        'pct_significant_transitions': '% Significant',
    }
    out = out.rename(columns=rename)
    out['Avg Growth (pp)'] = out['Avg Growth (pp)'].round(1)
    if '% Significant' in out.columns:
        out['% Significant'] = out['% Significant'].round(0)
    return out


def _build_equity_sheet(equity_rankings: pd.DataFrame, n: int = 25) -> pd.DataFrame:
    """Top equity schools across both subjects."""
    if equity_rankings.empty:
        return pd.DataFrame()

    frames = []
    for subj in ('ELA', 'Math'):
        sub = equity_rankings[equity_rankings['Subject'] == subj].copy()
        if 'avg_gap_change' in sub.columns:
            sub = sub.sort_values('avg_gap_change', ascending=False)
        sub = sub.head(n).reset_index(drop=True)
        # 'Subject' column already present in equity_rankings — no insert needed
        frames.append(sub)

    if not frames:
        return pd.DataFrame()

    out = pd.concat(frames, ignore_index=True)
    rename = {
        'equity_rank': 'Equity Rank',
        'School Name': 'School Name',
        'avg_gap_change': 'Avg Gap Change (pp)',
        'avg_growth_gap': 'Avg Growth Gap (pp)',
        'n_subgroups': 'No. Subgroups',
        'pct_narrowing': '% Transitions Narrowing',
    }
    out = out.rename(columns={k: v for k, v in rename.items() if k in out.columns})
    for col in ('Avg Gap Change (pp)', 'Avg Growth Gap (pp)', '% Transitions Narrowing'):
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors='coerce').round(1)
    return out


def _build_trends_sheet(trends: pd.DataFrame) -> pd.DataFrame:
    """Citywide avg proficiency by year × subject for All Students."""
    mask = trends['Student Group Value'].isin(ALL_STUDENTS_LABELS)
    city = trends[mask].groupby(['year', 'Subject'])['proficiency_pct'].mean().reset_index()
    city = city.sort_values(['Subject', 'year'])
    city['proficiency_pct'] = city['proficiency_pct'].round(1)
    city = city.rename(columns={
        'year': 'Year',
        'Subject': 'Subject',
        'proficiency_pct': 'Avg Proficiency (%)',
    })

    # Pivot to wide for readability
    pivot = city.pivot(index='Year', columns='Subject', values='Avg Proficiency (%)')
    pivot = pivot.reset_index()
    pivot.columns.name = None
    return pivot


def _build_directory(rankings: pd.DataFrame,
                     equity_rankings: pd.DataFrame) -> pd.DataFrame:
    """Combined school directory with growth + equity metrics for all schools."""
    frames = []
    for subj in ('ELA', 'Math'):
        gr = rankings[rankings['Subject'] == subj].copy() if not rankings.empty else pd.DataFrame()
        eq = equity_rankings[equity_rankings['Subject'] == subj].copy() if not equity_rankings.empty else pd.DataFrame()

        if gr.empty and eq.empty:
            continue

        if not gr.empty and not eq.empty:
            merged = gr.merge(
                eq[['School Name', 'equity_rank', 'avg_gap_change',
                    'pct_narrowing', 'n_subgroups']].rename(columns={
                    'equity_rank': 'Equity Rank',
                    'avg_gap_change': 'Avg Gap Change (pp)',
                    'pct_narrowing': '% Transitions Narrowing',
                    'n_subgroups': 'No. Equity Subgroups',
                }),
                on='School Name', how='outer'
            )
        elif not gr.empty:
            merged = gr.copy()
        else:
            merged = eq.copy()

        merged['Subject'] = subj
        frames.append(merged)

    if not frames:
        return pd.DataFrame()

    out = pd.concat(frames, ignore_index=True)
    out = out.sort_values(['Subject', 'rank'] if 'rank' in out.columns else ['Subject'])
    out = out.reset_index(drop=True)

    rename = {
        'rank': 'Growth Rank',
        'School Name': 'School Name',
        'avg_pp_growth': 'Avg Growth (pp)',
        'n_transitions': 'No. Transitions',
        'pct_significant_transitions': '% Significant',
    }
    out = out.rename(columns={k: v for k, v in rename.items() if k in out.columns})
    for col in ('Avg Growth (pp)', 'Avg Gap Change (pp)'):
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors='coerce').round(1)
    return out


# ═════════════════════════════════════════════════════════════════════════════
# Excel formatting helpers
# ═════════════════════════════════════════════════════════════════════════════

def _hex_to_argb(hex_color: str) -> str:
    """Convert 6-digit hex to openpyxl ARGB string."""
    return 'FF' + hex_color.upper()


def _apply_header_style(ws, row_idx: int, n_cols: int,
                        bg: str = HEADER_BG, fg: str = HEADER_FG,
                        bold: bool = True, font_size: int = 11) -> None:
    """Bold, coloured header row."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return

    fill = PatternFill(fill_type='solid', fgColor=_hex_to_argb(bg))
    font = Font(bold=bold, color=_hex_to_argb(fg), size=font_size)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _apply_alt_row_style(ws, start_row: int, n_data_rows: int,
                         n_cols: int, growth_col: int = None) -> None:
    """Alternating row shading + optional green/red colouring for a growth column."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return

    alt_fill = PatternFill(fill_type='solid', fgColor=_hex_to_argb(ALT_ROW_BG))
    pos_font = Font(color=_hex_to_argb(POSITIVE_FG), bold=True)
    neg_font = Font(color=_hex_to_argb(NEGATIVE_FG), bold=True)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    for r in range(start_row, start_row + n_data_rows):
        is_alt = (r - start_row) % 2 == 1
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            if is_alt:
                cell.fill = alt_fill
            align = left if c == 2 else center
            cell.alignment = align
            if growth_col and c == growth_col:
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.font = pos_font if val > 0 else neg_font


def _set_col_widths(ws, widths: list) -> None:
    """Set column widths by index (1-based)."""
    try:
        from openpyxl.utils import get_column_letter
    except ImportError:
        return
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _freeze_top_row(ws) -> None:
    ws.freeze_panes = ws['A2']


def _write_df_to_sheet(ws, df: pd.DataFrame,
                       start_row: int = 1,
                       growth_col_name: str = None) -> None:
    """Write a DataFrame to an openpyxl worksheet with formatting."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        # Fallback: no formatting
        for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        return

    n_cols = len(df.columns)

    # Write header
    for c_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=c_idx, value=str(col_name))
    _apply_header_style(ws, start_row, n_cols)

    # Find growth column index (1-based)
    growth_col = None
    if growth_col_name and growth_col_name in df.columns:
        growth_col = list(df.columns).index(growth_col_name) + 1

    # Write data rows
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    _apply_alt_row_style(ws, start_row + 1, len(df), n_cols, growth_col)


# ═════════════════════════════════════════════════════════════════════════════
# Executive Summary sheet (special layout)
# ═════════════════════════════════════════════════════════════════════════════

def _write_exec_summary(ws, exec_df: pd.DataFrame) -> None:
    """Write the executive summary with section banners and value cells."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        for r, row in enumerate(exec_df.itertuples(index=False), start=1):
            ws.cell(row=r, column=1, value=row.Metric)
            ws.cell(row=r, column=2, value=row.Value)
        return

    title_fill = PatternFill(fill_type='solid', fgColor=_hex_to_argb(TITLE_BG))
    title_font = Font(bold=True, color=_hex_to_argb(TITLE_FG), size=12)
    title_align = Alignment(horizontal='left', vertical='center')

    metric_font = Font(bold=False, size=11)
    value_font = Font(bold=True, size=11)
    metric_align = Alignment(horizontal='left', vertical='center', indent=2)
    value_align = Alignment(horizontal='right', vertical='center')

    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 30

    current_row = 1
    for _, row in exec_df.iterrows():
        row_type = str(row['Type'])
        metric = str(row['Metric'])
        value = str(row['Value'])

        if row_type == 'Section':
            # Section banner
            cell_a = ws.cell(row=current_row, column=1, value=metric)
            cell_b = ws.cell(row=current_row, column=2, value='')
            cell_a.fill = title_fill
            cell_a.font = title_font
            cell_a.alignment = title_align
            cell_b.fill = title_fill
        elif metric == '' and value == '':
            # Blank spacer row
            pass
        else:
            cell_a = ws.cell(row=current_row, column=1, value=metric)
            cell_b = ws.cell(row=current_row, column=2, value=value)
            cell_a.font = metric_font
            cell_a.alignment = metric_align
            cell_b.font = value_font
            cell_b.alignment = value_align

            # Alternating light background for non-section rows
            if (current_row % 2) == 0:
                try:
                    from openpyxl.styles import PatternFill
                    alt_fill = PatternFill(fill_type='solid',
                                          fgColor=_hex_to_argb(ALT_ROW_BG))
                    cell_a.fill = alt_fill
                    cell_b.fill = alt_fill
                except Exception:
                    pass

        current_row += 1

    ws.freeze_panes = None


def _build_geo_equity_sheet(geo_equity: pd.DataFrame) -> pd.DataFrame:
    """Geographic equity summary by DC quadrant."""
    if geo_equity.empty:
        return pd.DataFrame()

    rename = {
        'Quadrant': 'Quadrant',
        'Subject': 'Subject',
        'n_schools': 'No. Schools',
        'avg_pp_growth': 'Avg Cohort Growth (pp)',
        'median_pp_growth': 'Median Cohort Growth (pp)',
        'avg_baseline_proficiency': 'Avg Baseline Proficiency (%)',
        'avg_mean_proficiency': 'Avg Proficiency (%)',
        'gap_vs_nw_pp': 'Gap vs NW (pp)',
    }
    out = geo_equity.rename(columns={k: v for k, v in rename.items() if k in geo_equity.columns})
    # Ensure clean numeric columns
    for col in ('Avg Cohort Growth (pp)', 'Median Cohort Growth (pp)',
                'Avg Baseline Proficiency (%)', 'Avg Proficiency (%)', 'Gap vs NW (pp)'):
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors='coerce').round(1)
    quadrant_order = ['NE', 'NW', 'SE', 'SW']
    if 'Quadrant' in out.columns:
        out['Quadrant'] = pd.Categorical(out['Quadrant'], categories=quadrant_order, ordered=True)
        out = out.sort_values(['Subject', 'Quadrant']).reset_index(drop=True)
    return out


def _build_yoy_sheet(yoy_summary: pd.DataFrame) -> pd.DataFrame:
    """
    Same-grade year-over-year growth summary.

    Pivots the YoY summary to show All Students ELA and Math avg pp_change
    per school, sorted by avg_pp_change descending.
    """
    if yoy_summary.empty:
        return pd.DataFrame()

    all_students = yoy_summary[
        yoy_summary['Student Group Value'].isin({'All Students', 'All', 'Total'})
    ].copy()
    if all_students.empty:
        return pd.DataFrame()

    rename = {
        'School Name': 'School Name',
        'Subject': 'Subject',
        'n_transitions': 'No. Transitions',
        'avg_pp_change': 'Avg YoY Change (pp)',
        'median_pp_change': 'Median YoY Change (pp)',
        'pct_improving': '% Transition Years Improving',
        'max_pp_change': 'Best Year Change (pp)',
        'min_pp_change': 'Worst Year Change (pp)',
    }
    out = all_students.rename(columns={k: v for k, v in rename.items() if k in all_students.columns})
    for col in ('Avg YoY Change (pp)', 'Median YoY Change (pp)',
                'Best Year Change (pp)', 'Worst Year Change (pp)'):
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors='coerce').round(2)
    out = out.sort_values(['Subject', 'Avg YoY Change (pp)'], ascending=[True, False]).reset_index(drop=True)
    # Add rank within subject
    out.insert(0, 'Rank', out.groupby('Subject').cumcount() + 1)
    return out


def _build_covid_recovery_sheet(covid_recovery: pd.DataFrame) -> pd.DataFrame:
    """
    COVID Recovery summary sheet.

    Shows All Students summary per school × subject with pre-COVID baseline,
    COVID impact, recovery progress, net change, and recovery status.
    Sorted by Subject then net_vs_precovid_pp descending.
    """
    if covid_recovery.empty:
        return pd.DataFrame()

    rename = {
        'School Name': 'School Name',
        'Subject': 'Subject',
        'pct_2019': '2019 Proficiency (%)',
        'pct_2022': '2022 Proficiency (%)',
        'pct_2024': '2024 Proficiency (%)',
        'covid_impact_pp': 'COVID Impact 2019→2022 (pp)',
        'recovery_pp': 'Recovery 2022→2024 (pp)',
        'net_vs_precovid_pp': 'Net vs Pre-COVID (pp)',
        'recovery_status': 'Recovery Status',
    }
    out = covid_recovery.rename(columns={k: v for k, v in rename.items() if k in covid_recovery.columns})
    for col in ('2019 Proficiency (%)', '2022 Proficiency (%)', '2024 Proficiency (%)',
                'COVID Impact 2019→2022 (pp)', 'Recovery 2022→2024 (pp)', 'Net vs Pre-COVID (pp)'):
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors='coerce').round(2)
    out = out.sort_values(
        ['Subject', 'Net vs Pre-COVID (pp)'], ascending=[True, False]
    ).reset_index(drop=True)
    out.insert(0, 'Rank', out.groupby('Subject').cumcount() + 1)
    return out


def _build_trajectory_sheet(trajectory: pd.DataFrame) -> pd.DataFrame:
    """
    School Performance Trajectory sheet.

    Shows the long-run linear proficiency trend for every school × subject
    (All Students), including trend slope (pp/yr), R², trajectory class,
    and first/last proficiency.  Sorted by Subject then trend_slope_pp_yr
    descending.
    """
    if trajectory.empty:
        return pd.DataFrame()

    rename = {
        'School Name': 'School Name',
        'Subject': 'Subject',
        'n_years_with_data': 'Years With Data',
        'first_year': 'First Year',
        'last_year': 'Last Year',
        'avg_proficiency_pct': 'Avg Proficiency (%)',
        'first_proficiency_pct': 'First Year Proficiency (%)',
        'last_proficiency_pct': 'Last Year Proficiency (%)',
        'total_change_pp': 'Total Change (pp)',
        'trend_slope_pp_yr': 'Trend Slope (pp/yr)',
        'r_squared': 'R²',
        'trajectory_class': 'Trajectory Class',
    }
    out = trajectory.rename(columns={k: v for k, v in rename.items() if k in trajectory.columns})
    for col in ('Avg Proficiency (%)', 'First Year Proficiency (%)', 'Last Year Proficiency (%)',
                'Total Change (pp)', 'Trend Slope (pp/yr)', 'R²'):
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors='coerce').round(3)
    out = out.sort_values(
        ['Subject', 'Trend Slope (pp/yr)'], ascending=[True, False]
    ).reset_index(drop=True)
    out.insert(0, 'Rank', out.groupby('Subject').cumcount() + 1)
    return out


def _build_school_type_sheet(school_type: pd.DataFrame) -> pd.DataFrame:
    """
    School Type Performance sheet.

    Shows average proficiency, COVID recovery metrics, and cohort growth
    for each school type × subject combination.  Sorted by Subject then
    school type.
    """
    if school_type.empty:
        return pd.DataFrame()

    rename = {
        'School Type': 'School Type',
        'Subject': 'Subject',
        'n_schools': 'No. Schools',
        'avg_proficiency_pct': 'Avg Proficiency (%)',
        'covid_impact_pp': 'COVID Impact 2019→2022 (pp)',
        'recovery_pp': 'Recovery 2022→2024 (pp)',
        'net_vs_precovid_pp': 'Net vs Pre-COVID (pp)',
        'avg_cohort_growth_pp': 'Avg Cohort Growth (pp)',
    }
    out = school_type.rename(columns=rename)

    # Round numeric columns
    for col in ('Avg Proficiency (%)', 'COVID Impact 2019→2022 (pp)',
                'Recovery 2022→2024 (pp)', 'Net vs Pre-COVID (pp)',
                'Avg Cohort Growth (pp)'):
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors='coerce').round(2)

    type_order = [
        'Elementary', 'Middle School', 'High School',
        'Elementary-Middle', 'Middle-High',
    ]
    type_sort = {t: i for i, t in enumerate(type_order)}
    out['_type_sort'] = out['School Type'].map(type_sort).fillna(99)
    out = out.sort_values(['Subject', '_type_sort']).drop(columns=['_type_sort']).reset_index(drop=True)
    return out


def _build_grade_level_sheet(grade_level: pd.DataFrame) -> pd.DataFrame:
    """
    Grade-Level Performance sheet.

    Shows average proficiency, COVID impact, recovery, net change, and
    average YoY growth for every grade × subject combination.  Sorted by
    Subject then grade order (Grade 3 → HS).
    """
    if grade_level.empty:
        return pd.DataFrame()

    rename = {
        'grade': 'Grade',
        'Subject': 'Subject',
        'n_schools': 'No. Schools',
        'avg_proficiency_pct': 'Avg Proficiency (%)',
        'covid_impact_pp': 'COVID Impact 2019→2022 (pp)',
        'recovery_pp': 'Recovery 2022→2024 (pp)',
        'net_vs_precovid_pp': 'Net vs Pre-COVID (pp)',
        'avg_yoy_growth_pp': 'Avg YoY Growth (pp)',
    }
    out = grade_level.rename(columns=rename)

    # Round numeric columns
    for col in ('Avg Proficiency (%)', 'COVID Impact 2019→2022 (pp)',
                'Recovery 2022→2024 (pp)', 'Net vs Pre-COVID (pp)',
                'Avg YoY Growth (pp)'):
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors='coerce').round(2)

    grade_order = ['Grade 3', 'Grade 4', 'Grade 5', 'Grade 6', 'Grade 7', 'Grade 8', 'HS']
    grade_sort = {g: i for i, g in enumerate(grade_order)}
    out['_grade_sort'] = out['Grade'].map(grade_sort).fillna(99)
    out = out.sort_values(['Subject', '_grade_sort']).drop(columns=['_grade_sort']).reset_index(drop=True)
    return out


def _build_subgroup_sheet(subgroup_summary: pd.DataFrame) -> pd.DataFrame:
    """
    Subgroup Proficiency Trend sheet.

    Shows average proficiency, COVID impact, recovery, net change, and
    average YoY growth for every student subgroup × subject combination.
    Sorted by Subject then subgroup display order.
    """
    if subgroup_summary.empty:
        return pd.DataFrame()

    rename = {
        'subgroup': 'Student Subgroup',
        'Subject': 'Subject',
        'n_schools': 'No. Schools',
        'avg_proficiency_pct': 'Avg Proficiency (%)',
        'covid_impact_pp': 'COVID Impact 2019→2022 (pp)',
        'recovery_pp': 'Recovery 2022→2024 (pp)',
        'net_vs_precovid_pp': 'Net vs Pre-COVID (pp)',
        'avg_yoy_growth_pp': 'Avg YoY Growth (pp)',
    }
    out = subgroup_summary.rename(columns=rename)

    # Round numeric columns
    for col in ('Avg Proficiency (%)', 'COVID Impact 2019→2022 (pp)',
                'Recovery 2022→2024 (pp)', 'Net vs Pre-COVID (pp)',
                'Avg YoY Growth (pp)'):
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors='coerce').round(2)

    # NOTE: "Econ Dis" and "EL Active" are the abbreviated OSSE source labels
    # that appear verbatim in combined_all_years.csv / subgroup_summary.csv.
    subgroup_order = [
        'All Students', 'Male', 'Female',
        'Black or African American', 'Hispanic/Latino of any race',
        'White', 'Asian', 'Two or more races',
        'Econ Dis', 'EL Active', 'Students with Disabilities',
    ]
    sg_sort = {s: i for i, s in enumerate(subgroup_order)}
    out['_sg_sort'] = out['Student Subgroup'].map(sg_sort).fillna(99)
    out = out.sort_values(['Subject', '_sg_sort']).drop(columns=['_sg_sort']).reset_index(drop=True)
    return out


def _build_consistency_sheet(consistency: pd.DataFrame) -> pd.DataFrame:
    """
    School Performance Consistency sheet.

    Shows year-to-year performance stability for every school × subject
    (All Students), including std deviation, CV, range, and consistency class.
    Sorted by Subject then consistency class then avg proficiency descending.
    """
    if consistency.empty:
        return pd.DataFrame()

    rename = {
        'School Name': 'School Name',
        'Subject': 'Subject',
        'consistency_class': 'Consistency Class',
        'n_years_with_data': 'Years With Data',
        'first_year': 'First Year',
        'last_year': 'Last Year',
        'avg_proficiency_pct': 'Avg Proficiency (%)',
        'std_proficiency_pct': 'Std Dev (pp)',
        'cv_proficiency_pct': 'CV (%)',
        'min_proficiency_pct': 'Min Proficiency (%)',
        'max_proficiency_pct': 'Max Proficiency (%)',
        'range_proficiency_pp': 'Range (max−min, pp)',
    }
    out = consistency.rename(columns={k: v for k, v in rename.items() if k in consistency.columns})

    for col in ('Avg Proficiency (%)', 'Std Dev (pp)', 'CV (%)',
                'Min Proficiency (%)', 'Max Proficiency (%)', 'Range (max−min, pp)'):
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors='coerce').round(2)

    class_order = [
        'High-Consistent', 'High-Volatile',
        'Low-Consistent', 'Low-Volatile',
        'Insufficient Data',
    ]
    class_sort = {c: i for i, c in enumerate(class_order)}
    out['_cls_sort'] = out['Consistency Class'].map(class_sort).fillna(99)
    out = out.sort_values(
        ['Subject', '_cls_sort', 'Avg Proficiency (%)'],
        ascending=[True, True, False],
    ).drop(columns=['_cls_sort']).reset_index(drop=True)
    out.insert(0, 'Rank', out.groupby('Subject').cumcount() + 1)
    return out


def _build_performance_index_sheet(performance_index: pd.DataFrame) -> pd.DataFrame:
    """
    Multi-Metric School Performance Index sheet.

    Shows the composite score, quintile, and four component percentile-rank
    scores for every school × subject (All Students).  Sorted by Subject then
    quintile then composite score descending.
    """
    if performance_index.empty:
        return pd.DataFrame()

    rename = {
        'School Name': 'School Name',
        'Subject': 'Subject',
        'composite_score': 'Composite Score (0–100)',
        'composite_quintile': 'Quintile',
        'n_components': 'No. Components',
        'proficiency_score': 'Proficiency Score',
        'growth_score': 'Growth Score',
        'recovery_score': 'Recovery Score',
        'trajectory_score': 'Trajectory Score',
        'proficiency_pct': 'Avg Proficiency (%)',
        'cohort_growth_pp': 'Avg Cohort Growth (pp)',
        'covid_recovery_pp': 'COVID Recovery (pp)',
        'trajectory_slope_pp_yr': 'Trend Slope (pp/yr)',
    }
    out = performance_index.rename(
        columns={k: v for k, v in rename.items() if k in performance_index.columns}
    )

    for col in ('Composite Score (0–100)', 'Proficiency Score', 'Growth Score',
                'Recovery Score', 'Trajectory Score', 'Avg Proficiency (%)',
                'Avg Cohort Growth (pp)', 'COVID Recovery (pp)', 'Trend Slope (pp/yr)'):
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors='coerce').round(2)

    quintile_order = [
        'Q5 – Top Performers', 'Q4 – Above Average', 'Q3 – Middle',
        'Q2 – Below Average', 'Q1 – Bottom Performers', 'Insufficient Data',
    ]
    q_sort = {q: i for i, q in enumerate(quintile_order)}
    out['_q_sort'] = out['Quintile'].map(q_sort).fillna(99)
    out = out.sort_values(
        ['Subject', '_q_sort', 'Composite Score (0–100)'],
        ascending=[True, True, False],
    ).drop(columns=['_q_sort']).reset_index(drop=True)
    out.insert(0, 'Rank', out.groupby('Subject').cumcount() + 1)
    return out


# ═════════════════════════════════════════════════════════════════════════════
# Main
# ═════════════════════════════════════════════════════════════════════════════

def main() -> None:
    print('=' * 70)
    print('GENERATING SUMMARY REPORT — DC Schools Test Score Analysis')
    print('=' * 70)
    print()

    # ── Load inputs ────────────────────────────────────────────────────────
    cohort_summary = _load(COHORT_SUMMARY_FILE, 'cohort_growth_summary.csv')
    rankings = _load(RANKINGS_FILE, 'school_rankings.csv')
    equity_rankings = _load(EQUITY_RANKINGS_FILE, 'school_equity_rankings.csv')
    equity_summary = _load(EQUITY_SUMMARY_FILE, 'equity_gap_summary.csv')
    trends = _load(TRENDS_FILE, 'proficiency_trends.csv')
    # Geographic equity is optional — generated by geographic_equity_analysis.py
    geo_equity = pd.DataFrame()
    if os.path.isfile(GEO_EQUITY_FILE):
        geo_equity = pd.read_csv(GEO_EQUITY_FILE)
        print(f'  Loaded geographic_equity_by_quadrant.csv: {len(geo_equity):,} rows')
    else:
        print(f'  NOTE: {GEO_EQUITY_FILE} not found — Sheet 7 will be skipped.')
    # YoY growth summary is optional — generated by yoy_growth_analysis.py
    yoy_summary = pd.DataFrame()
    if os.path.isfile(YOY_SUMMARY_FILE):
        yoy_summary = pd.read_csv(YOY_SUMMARY_FILE)
        print(f'  Loaded yoy_growth_summary.csv: {len(yoy_summary):,} rows')
    else:
        print(f'  NOTE: {YOY_SUMMARY_FILE} not found — Sheet 8 will be skipped.')
    # COVID recovery summary is optional — generated by covid_recovery_analysis.py
    covid_recovery = pd.DataFrame()
    if os.path.isfile(COVID_RECOVERY_FILE):
        covid_recovery = pd.read_csv(COVID_RECOVERY_FILE)
        print(f'  Loaded covid_recovery_summary.csv: {len(covid_recovery):,} rows')
    else:
        print(f'  NOTE: {COVID_RECOVERY_FILE} not found — Sheet 9 will be skipped.')
    # School trajectory classification is optional — generated by school_trajectory_analysis.py
    trajectory = pd.DataFrame()
    if os.path.isfile(TRAJECTORY_FILE):
        trajectory = pd.read_csv(TRAJECTORY_FILE)
        print(f'  Loaded school_trajectory_classification.csv: {len(trajectory):,} rows')
    else:
        print(f'  NOTE: {TRAJECTORY_FILE} not found — Sheet 10 will be skipped.')
    # School type summary is optional — generated by school_type_analysis.py
    school_type = pd.DataFrame()
    if os.path.isfile(SCHOOL_TYPE_FILE):
        school_type = pd.read_csv(SCHOOL_TYPE_FILE)
        print(f'  Loaded school_type_summary.csv: {len(school_type):,} rows')
    else:
        print(f'  NOTE: {SCHOOL_TYPE_FILE} not found — Sheet 11 will be skipped.')
    # Grade-level summary is optional — generated by grade_level_analysis.py
    grade_level = pd.DataFrame()
    if os.path.isfile(GRADE_LEVEL_FILE):
        grade_level = pd.read_csv(GRADE_LEVEL_FILE)
        print(f'  Loaded grade_level_summary.csv: {len(grade_level):,} rows')
    else:
        print(f'  NOTE: {GRADE_LEVEL_FILE} not found — Sheet 12 will be skipped.')
    # Subgroup summary is optional — generated by subgroup_trend_analysis.py
    subgroup_summary = pd.DataFrame()
    if os.path.isfile(SUBGROUP_SUMMARY_FILE):
        subgroup_summary = pd.read_csv(SUBGROUP_SUMMARY_FILE)
        print(f'  Loaded subgroup_summary.csv: {len(subgroup_summary):,} rows')
    else:
        print(f'  NOTE: {SUBGROUP_SUMMARY_FILE} not found — Sheet 13 will be skipped.')
    # School consistency is optional — generated by school_consistency_analysis.py
    consistency = pd.DataFrame()
    if os.path.isfile(CONSISTENCY_FILE):
        consistency = pd.read_csv(CONSISTENCY_FILE)
        print(f'  Loaded school_consistency.csv: {len(consistency):,} rows')
    else:
        print(f'  NOTE: {CONSISTENCY_FILE} not found — Sheet 14 will be skipped.')
    # Performance index is optional — generated by school_performance_index.py
    performance_index = pd.DataFrame()
    if os.path.isfile(PERFORMANCE_INDEX_FILE):
        performance_index = pd.read_csv(PERFORMANCE_INDEX_FILE)
        print(f'  Loaded school_performance_index.csv: {len(performance_index):,} rows')
    else:
        print(f'  NOTE: {PERFORMANCE_INDEX_FILE} not found — Sheet 15 will be skipped.')
    print()

    # ── Build sheet data ───────────────────────────────────────────────────
    exec_df = _build_exec_summary(cohort_summary, rankings, equity_rankings, trends)
    top_ela = _build_top_growth(rankings, 'ELA', TOP_N)
    top_math = _build_top_growth(rankings, 'Math', TOP_N)
    equity_sheet = _build_equity_sheet(equity_rankings, TOP_N)
    trends_sheet = _build_trends_sheet(trends)
    directory = _build_directory(rankings, equity_rankings)
    geo_equity_sheet = _build_geo_equity_sheet(geo_equity)
    yoy_sheet = _build_yoy_sheet(yoy_summary)
    covid_sheet = _build_covid_recovery_sheet(covid_recovery)
    trajectory_sheet = _build_trajectory_sheet(trajectory)
    school_type_sheet = _build_school_type_sheet(school_type)
    grade_level_sheet = _build_grade_level_sheet(grade_level)
    subgroup_sheet = _build_subgroup_sheet(subgroup_summary)
    consistency_sheet = _build_consistency_sheet(consistency)
    performance_index_sheet = _build_performance_index_sheet(performance_index)

    # ── Write workbook ─────────────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        _has_openpyxl = True
    except ImportError:
        _has_openpyxl = False
        print('WARNING: openpyxl not installed — writing plain CSV fallback.')

    if _has_openpyxl:
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Sheet 1 — Executive Summary
        ws1 = wb.create_sheet('Executive Summary')
        _write_exec_summary(ws1, exec_df)
        print('  ✓ Sheet 1: Executive Summary')

        # Sheet 2 — Top Growth ELA
        ws2 = wb.create_sheet('Top Growth (ELA)')
        _write_df_to_sheet(ws2, top_ela, growth_col_name='Avg Growth (pp)')
        _set_col_widths(ws2, [8, 45, 16, 16, 14])
        _freeze_top_row(ws2)
        print(f'  ✓ Sheet 2: Top Growth (ELA) — {len(top_ela)} schools')

        # Sheet 3 — Top Growth Math
        ws3 = wb.create_sheet('Top Growth (Math)')
        _write_df_to_sheet(ws3, top_math, growth_col_name='Avg Growth (pp)')
        _set_col_widths(ws3, [8, 45, 16, 16, 14])
        _freeze_top_row(ws3)
        print(f'  ✓ Sheet 3: Top Growth (Math) — {len(top_math)} schools')

        # Sheet 4 — Top Equity Schools
        ws4 = wb.create_sheet('Top Equity Schools')
        if not equity_sheet.empty:
            _write_df_to_sheet(ws4, equity_sheet, growth_col_name='Avg Gap Change (pp)')
            _set_col_widths(ws4, [8, 8, 45, 10, 20, 20, 20, 24])
            _freeze_top_row(ws4)
        print(f'  ✓ Sheet 4: Top Equity Schools — {len(equity_sheet)} rows')

        # Sheet 5 — Proficiency Trends
        ws5 = wb.create_sheet('Proficiency Trends')
        _write_df_to_sheet(ws5, trends_sheet)
        _set_col_widths(ws5, [8, 18, 18])
        _freeze_top_row(ws5)
        print(f'  ✓ Sheet 5: Proficiency Trends — {len(trends_sheet)} year rows')

        # Sheet 6 — School Directory
        ws6 = wb.create_sheet('School Directory')
        if not directory.empty:
            _write_df_to_sheet(ws6, directory, growth_col_name='Avg Growth (pp)')
            _set_col_widths(ws6, [8, 12, 45, 16, 16, 14, 12, 20, 20, 24])
            _freeze_top_row(ws6)
        print(f'  ✓ Sheet 6: School Directory — {len(directory)} rows')

        # Sheet 7 — Geographic Equity (optional, requires geographic_equity_analysis.py)
        if not geo_equity_sheet.empty:
            ws7 = wb.create_sheet('Geographic Equity')
            _write_df_to_sheet(ws7, geo_equity_sheet, growth_col_name='Avg Cohort Growth (pp)')
            _set_col_widths(ws7, [10, 10, 14, 22, 24, 26, 22, 18])
            _freeze_top_row(ws7)
            print(f'  ✓ Sheet 7: Geographic Equity — {len(geo_equity_sheet)} rows')
        else:
            print('  – Sheet 7: Geographic Equity skipped (run geographic_equity_analysis.py first)')

        # Sheet 8 — YoY Growth (optional, requires yoy_growth_analysis.py)
        if not yoy_sheet.empty:
            ws8 = wb.create_sheet('YoY Growth')
            _write_df_to_sheet(ws8, yoy_sheet, growth_col_name='Avg YoY Change (pp)')
            _set_col_widths(ws8, [8, 10, 45, 16, 22, 24, 28, 22, 22])
            _freeze_top_row(ws8)
            print(f'  ✓ Sheet 8: YoY Growth — {len(yoy_sheet)} rows')
        else:
            print('  – Sheet 8: YoY Growth skipped (run yoy_growth_analysis.py first)')

        # Sheet 9 — COVID Recovery (optional, requires covid_recovery_analysis.py)
        if not covid_sheet.empty:
            ws9 = wb.create_sheet('COVID Recovery')
            _write_df_to_sheet(ws9, covid_sheet, growth_col_name='Net vs Pre-COVID (pp)')
            _set_col_widths(ws9, [8, 10, 45, 18, 18, 18, 26, 24, 22, 22])
            _freeze_top_row(ws9)
            print(f'  ✓ Sheet 9: COVID Recovery — {len(covid_sheet)} rows')
        else:
            print('  – Sheet 9: COVID Recovery skipped (run covid_recovery_analysis.py first)')

        # Sheet 10 — School Trajectories (optional, requires school_trajectory_analysis.py)
        if not trajectory_sheet.empty:
            ws10 = wb.create_sheet('School Trajectories')
            _write_df_to_sheet(ws10, trajectory_sheet, growth_col_name='Trend Slope (pp/yr)')
            _set_col_widths(ws10, [8, 10, 45, 14, 12, 12, 20, 22, 22, 18, 18, 8, 22])
            _freeze_top_row(ws10)
            print(f'  ✓ Sheet 10: School Trajectories — {len(trajectory_sheet)} rows')
        else:
            print('  – Sheet 10: School Trajectories skipped (run school_trajectory_analysis.py first)')

        # Sheet 11 — School Types (optional, requires school_type_analysis.py)
        if not school_type_sheet.empty:
            ws11 = wb.create_sheet('School Types')
            _write_df_to_sheet(ws11, school_type_sheet, growth_col_name='Avg Cohort Growth (pp)')
            _set_col_widths(ws11, [22, 10, 12, 20, 28, 28, 24, 26])
            _freeze_top_row(ws11)
            print(f'  ✓ Sheet 11: School Types — {len(school_type_sheet)} rows')
        else:
            print('  – Sheet 11: School Types skipped (run school_type_analysis.py first)')

        # Sheet 12 — Grade Levels (optional, requires grade_level_analysis.py)
        if not grade_level_sheet.empty:
            ws12 = wb.create_sheet('Grade Levels')
            _write_df_to_sheet(ws12, grade_level_sheet, growth_col_name='Avg YoY Growth (pp)')
            _set_col_widths(ws12, [12, 10, 12, 20, 28, 24, 24, 22])
            _freeze_top_row(ws12)
            print(f'  ✓ Sheet 12: Grade Levels — {len(grade_level_sheet)} rows')
        else:
            print('  – Sheet 12: Grade Levels skipped (run grade_level_analysis.py first)')

        # Sheet 13 — Subgroups (optional, requires subgroup_trend_analysis.py)
        if not subgroup_sheet.empty:
            ws13 = wb.create_sheet('Subgroups')
            _write_df_to_sheet(ws13, subgroup_sheet, growth_col_name='Avg YoY Growth (pp)')
            _set_col_widths(ws13, [30, 10, 12, 20, 28, 24, 24, 22])
            _freeze_top_row(ws13)
            print(f'  ✓ Sheet 13: Subgroups — {len(subgroup_sheet)} rows')
        else:
            print('  – Sheet 13: Subgroups skipped (run subgroup_trend_analysis.py first)')

        # Sheet 14 — Consistency (optional, requires school_consistency_analysis.py)
        if not consistency_sheet.empty:
            ws14 = wb.create_sheet('Consistency')
            _write_df_to_sheet(ws14, consistency_sheet)
            _set_col_widths(ws14, [8, 10, 45, 22, 14, 12, 12, 20, 12, 12, 20, 20, 22])
            _freeze_top_row(ws14)
            print(f'  ✓ Sheet 14: Consistency — {len(consistency_sheet)} rows')
        else:
            print('  – Sheet 14: Consistency skipped (run school_consistency_analysis.py first)')

        # Sheet 15 — Performance Index (optional, requires school_performance_index.py)
        if not performance_index_sheet.empty:
            ws15 = wb.create_sheet('Performance Index')
            _write_df_to_sheet(ws15, performance_index_sheet,
                               growth_col_name='Composite Score (0–100)')
            _set_col_widths(ws15, [8, 10, 45, 18, 18, 16, 16, 16, 16, 16, 20, 22, 22, 18])
            _freeze_top_row(ws15)
            print(f'  ✓ Sheet 15: Performance Index — {len(performance_index_sheet)} rows')
        else:
            print('  – Sheet 15: Performance Index skipped '
                  '(run school_performance_index.py first)')

        wb.save(REPORT_FILE)
        print(f'\n✓ Saved: {REPORT_FILE}')

    else:
        # Fallback: write CSVs if openpyxl is not available
        fallback_base = REPORT_FILE.replace('.xlsx', '')
        exec_df[['Metric', 'Value']].to_csv(fallback_base + '_exec.csv', index=False)
        top_ela.to_csv(fallback_base + '_top_ela.csv', index=False)
        top_math.to_csv(fallback_base + '_top_math.csv', index=False)
        print(f'  Fallback CSV files written with prefix {fallback_base}_*')

    print()
    print('=' * 70)
    print('SUMMARY REPORT COMPLETE!')
    print('=' * 70)


if __name__ == '__main__':
    main()
