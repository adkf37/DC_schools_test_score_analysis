"""
Load and convert wide-format OSSE school-level assessment files into the
standard combined_all_years.csv format consumed by analyze_cohort_growth.py.

This loader handles the "School and Demographic Group Aggregation" Excel files
found in input_data/School and Demographic Group Aggregation/.  Each workbook
follows the same layout:
  - Multiple sheets (Overall, Female, Male, AmIndAlNa, Asian, …)
  - Rows 0-4 are metadata / blank; row 5 (0-indexed) is the header
  - Each data row = one school × one grade
  - ELA and Math columns appear side-by-side in the same row

Output: output_data/combined_all_years.csv (same schema as load_clean_data.py)

Usage:
    python src/load_wide_format_data.py
"""
import os
import re
import sys
import openpyxl
import pandas as pd
import numpy as np

# ── Paths ─────────────────────────────────────────────────────────────────────
CURRENT_PATH = os.path.abspath(os.path.dirname(__file__))
REPO_ROOT = os.path.abspath(os.path.join(CURRENT_PATH, ".."))
INPUT_PATH = os.path.join(REPO_ROOT, "input_data")
OUTPUT_PATH = os.path.join(REPO_ROOT, "output_data")

# ── File → year / assessment mapping ──────────────────────────────────────────
# Each tuple: (filename_pattern_keywords, year, assessment_name)
# Patterns are matched case-insensitively against file basenames.
# Historical files (2015-16 through 2018-19) are in the same directory and
# share the same wide-format schema; they extend cohort-tracking coverage
# back to 2016 (decision D-020).
FILE_YEAR_MAP = [
    (["2015-16"],              2016, "PARCC"),
    (["2016-17"],              2017, "PARCC"),
    (["2017-18"],              2018, "PARCC"),
    (["2018-19"],              2019, "PARCC"),
    (["2021-22"],              2022, "PARCC"),
    (["2022-23"],              2023, "PARCC"),
    (["2023-2024", "2023-24"], 2024, "DCCAPE"),
]

# ── Demographic sheet → (Student Group, Student Group Value) ──────────────────
# Current (2021-22 / 2022-23 / 2023-24) files use short abbreviation sheet names.
# Historical (2015-16 through 2018-19) files use longer descriptive names that
# evolved across years.  All variants are mapped to the same standard labels so
# analyze_cohort_growth.py can track cohorts across the full date range.
SHEET_SUBGROUP = {
    # ── Current abbreviated sheet names (2021-22, 2022-23, 2023-24) ──────────
    "Overall":   ("All Students", "All Students"),
    "Female":    ("Gender", "Female"),
    "Male":      ("Gender", "Male"),
    "AmIndAlNa": ("Race/Ethnicity", "American Indian/Alaska Native"),
    "Asian":     ("Race/Ethnicity", "Asian"),
    "BlAfAm":    ("Race/Ethnicity", "Black or African American"),
    "HisLa":     ("Race/Ethnicity", "Hispanic/Latino of any race"),
    "Multi":     ("Race/Ethnicity", "Two or more races"),
    "PacIslNaH": ("Race/Ethnicity", "Pacific Islander/Native Hawaiian"),
    "WhCau":     ("Race/Ethnicity", "White"),
    "EL":        ("English Learner Status", "EL Active"),
    "SpEd":      ("Special Education Status", "Students with Disabilities"),
    "EconDisad": ("Economic Status", "Econ Dis"),
    # AtRisk (2021-22 only) maps to the same label as EconDisad because
    # analyze_cohort_growth.py explicitly treats them as close proxies for
    # cross-year comparison (see decision D-002 and subgroup_map therein).
    # These sheets never co-exist in the same year file, so no duplicate rows
    # are produced.
    "AtRisk":    ("Economic Status", "Econ Dis"),

    # ── Historical sheet names: 2015-16 ──────────────────────────────────────
    # 2015-16 uses "All" instead of "Overall" and longer descriptive names.
    "All":                             ("All Students", "All Students"),
    "Special Education Students":      ("Special Education Status", "Students with Disabilities"),
    "Eng Lang Learner Students":       ("English Learner Status", "EL Active"),
    "Free-Red Price Meal Students":    ("Economic Status", "Econ Dis"),

    # ── Historical sheet names: 2016-17 and 2017-18 ──────────────────────────
    # These files append " Students" to each demographic group name.
    "Female Students":                 ("Gender", "Female"),
    "Male Students":                   ("Gender", "Male"),
    "American Indian Students":        ("Race/Ethnicity", "American Indian/Alaska Native"),
    "Asian Students":                  ("Race/Ethnicity", "Asian"),
    "Black Students":                  ("Race/Ethnicity", "Black or African American"),
    "Hispanic Students":               ("Race/Ethnicity", "Hispanic/Latino of any race"),
    # 2017-18 file has a capitalization typo in the sheet name
    "HIspanic Students":               ("Race/Ethnicity", "Hispanic/Latino of any race"),
    "Multiracial Students":            ("Race/Ethnicity", "Two or more races"),
    "Nat Hawaiian-Pac Islnd Students": ("Race/Ethnicity", "Pacific Islander/Native Hawaiian"),
    "White Students":                  ("Race/Ethnicity", "White"),
    "ELL Students (Active & Monitor)": ("English Learner Status", "EL Active"),
    "EL Students (Active & Monitor)":  ("English Learner Status", "EL Active"),
    "SPED Students (Active & Monitor)": ("Special Education Status", "Students with Disabilities"),
    "Econ Disadvantaged Students":     ("Economic Status", "Econ Dis"),
    "At-Risk Students":                ("Economic Status", "Econ Dis"),

    # ── Historical sheet names: 2018-19 ──────────────────────────────────────
    # 2018-19 drops " Students" suffix but is otherwise similar.
    "American Indian":                 ("Race/Ethnicity", "American Indian/Alaska Native"),
    "Black":                           ("Race/Ethnicity", "Black or African American"),
    "Hispanic":                        ("Race/Ethnicity", "Hispanic/Latino of any race"),
    "Multiracial":                     ("Race/Ethnicity", "Two or more races"),
    "Pac Islander-Nat Hawaiian":       ("Race/Ethnicity", "Pacific Islander/Native Hawaiian"),
    "White":                           ("Race/Ethnicity", "White"),
    "EL (Active and Monitored)":       ("English Learner Status", "EL Active"),
    "SWD (Active)":                    ("Special Education Status", "Students with Disabilities"),
    "At-Risk":                         ("Economic Status", "Econ Dis"),
}

# Grades to exclude (school-wide totals rather than a specific cohort grade)
SKIP_GRADES = {"ALL", "ALL GRADES"}

# ── Column name patterns for dynamic header lookup ────────────────────────────
# Each value is a list of substrings to search for (case-insensitive) in the header.
# The first matching column index is used.
# Patterns cover both the current (2021+) column names and the older (2015-19)
# column name variants (decision D-020).
COL_PATTERNS = {
    "school_code":   ["school code"],
    "school_name":   ["school name"],
    "grade":         ["grade"],
    # PARCC/DCCAPE ELA — all years use the same column names
    "ela_total":     ["# of english language arts", "# of ela"],
    "ela_prof":      ["ela # proficient", "ela # prof"],
    "ela_pct":       ["ela - % proficient", "ela % proficient"],
    # PARCC/DCCAPE Math — all years use the same column names
    "math_total":    ["# math test takers"],
    "math_prof":     ["math - # proficient", "math # proficient"],
    "math_pct":      ["math - % proficient", "math % proficient"],
    # MSAA ELA — 2018-19+ uses "msaa - …"; 2015-16/2016-17 uses "msaa ela …"
    "msaa_ela_total": ["msaa - # of english language arts", "msaa - # of ela",
                       "msaa ela # of test takers"],
    "msaa_ela_prof":  ["msaa - ela # proficient", "msaa ela - # proficient"],
    "msaa_ela_pct":   ["msaa - ela - % proficient", "msaa - ela % proficient",
                       "msaa ela - % proficient"],
    # MSAA Math — same naming split as ELA
    "msaa_math_total": ["msaa - # of math", "msaa math - # test takers"],
    "msaa_math_prof":  ["msaa - math # proficient", "msaa math - # proficient"],
    "msaa_math_pct":   ["msaa - math - % proficient", "msaa - math % proficient",
                        "msaa math - % proficient"],
}


def _find_col_indices(header: tuple) -> dict:
    """
    Return a dict mapping field name → column index by searching header for
    known column name substrings.  Returns -1 for any column not found.
    """
    indices = {}
    # Normalise header entries (strip spaces, lower-case)
    norm = [str(h).strip().lower() if h is not None else "" for h in header]
    for field, patterns in COL_PATTERNS.items():
        idx = -1
        for pat in patterns:
            for i, h in enumerate(norm):
                if pat.lower() in h:
                    idx = i
                    break
            if idx >= 0:
                break
        indices[field] = idx
    # Grade column: use the *first* occurrence of 'grade' that is NOT 'school name'
    # (The 'school name' cell often contains 'grade' nowhere, so this is fine.)
    indices["school_code"] = norm.index("school code") if "school code" in norm else 0
    indices["school_name"] = next(
        (i for i, h in enumerate(norm) if "school name" in h), 1
    )
    indices["grade"] = next(
        (i for i, h in enumerate(norm) if h == "grade"), 2
    )
    return indices


# ═════════════════════════════════════════════════════════════════════════════
# Helpers
# ═════════════════════════════════════════════════════════════════════════════

def _coerce(val):
    """Return a numeric float or NaN from a raw cell value."""
    if val is None:
        return np.nan
    s = str(val).strip().upper()
    # Consolidated suppression check: empty strings, OSSE suppression codes,
    # and any value starting with N<, <=, or >= (e.g. "n < 10", "<5%", ">95%")
    if s in ("", "DS", "N/A", "NA", "."):
        return np.nan
    if re.match(r"^(N\s*<|<[^-]|>|\.)|\bDS\b", s):
        return np.nan
    s = s.replace("%", "").replace(",", "").strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return np.nan


def normalize_grade(raw: str) -> str:
    """'03' → 'Grade 3',  '11' → 'Grade 11',  'ALL' → 'All'."""
    s = str(raw).strip()
    if s.upper() in ("ALL", "ALL GRADES"):
        return "All"
    m = re.match(r"^(\d{1,2})$", s)
    if m:
        return f"Grade {int(m.group(1))}"
    # Already like 'Grade 8'
    return s


def find_files() -> list:
    """
    Walk input_data/ recursively and return a list of
    (filepath, year, assessment_name) for files that match known year patterns.
    """
    found = []
    for dirpath, _dirs, filenames in os.walk(INPUT_PATH):
        for fname in filenames:
            if not fname.lower().endswith(".xlsx"):
                continue
            # Skip temporary Excel files
            if fname.startswith("~$"):
                continue
            fpath = os.path.join(dirpath, fname)
            fname_lower = fname.lower()
            for keywords, year, assessment in FILE_YEAR_MAP:
                if any(kw.lower() in fname_lower for kw in keywords):
                    found.append((fpath, year, assessment))
                    break
    # Deduplicate by year (keep first match)
    seen_years = set()
    deduped = []
    for item in found:
        if item[1] not in seen_years:
            deduped.append(item)
            seen_years.add(item[1])
    deduped.sort(key=lambda x: x[1])
    return deduped


def read_wide_sheet(wb, sheet_name: str, year: int, assessment: str,
                    student_group: str, student_group_value: str) -> pd.DataFrame:
    """
    Read one demographic sheet and return a long-format DataFrame with one row
    per school × grade × subject.
    """
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)

    # Skip the first 5 rows (metadata / blank)
    for _ in range(5):
        try:
            next(rows_iter)
        except StopIteration:
            return pd.DataFrame()

    # Row 6 (0-index 5) is the header — consume it
    try:
        header = next(rows_iter)
    except StopIteration:
        return pd.DataFrame()

    # Dynamically locate column indices from the header
    col_idx = _find_col_indices(header)

    records = []
    for row in rows_iter:
        school_code_val = row[col_idx["school_code"]] if col_idx["school_code"] >= 0 else None
        if school_code_val is None:
            continue
        school_code = str(school_code_val).strip()
        school_name = (str(row[col_idx["school_name"]]).strip()
                       if col_idx["school_name"] >= 0 and row[col_idx["school_name"]] is not None
                       else "")
        grade_raw   = (str(row[col_idx["grade"]]).strip().upper()
                       if col_idx["grade"] >= 0 and row[col_idx["grade"]] is not None
                       else "")

        # Skip school-total rows (grade = 'ALL')
        if grade_raw in SKIP_GRADES:
            continue

        grade_label = normalize_grade(str(row[col_idx["grade"]]).strip()
                                      if col_idx["grade"] >= 0 else "")

        # Helper: safe cell access
        def cell(key):
            idx = col_idx.get(key, -1)
            return row[idx] if idx >= 0 and idx < len(row) else None

        # Helper to build one subject record
        def make_record(subject, total_key, prof_key, pct_key, assess_name):
            total = _coerce(cell(total_key))
            prof  = _coerce(cell(prof_key))
            pct_f = _coerce(cell(pct_key))
            # Convert fraction → percentage (skip if already looks like 0-100 range)
            if not np.isnan(pct_f) and pct_f <= 1.0:
                pct = round(pct_f * 100, 6)
            else:
                pct = pct_f
            return {
                "Aggregation Level":   "School",
                # LEA Code/Name are not present in wide-format files. Placeholders
                # ("0" / "DC Schools") are used so pandas does not coerce the
                # column to float64 NaN, which would cause groupby in
                # analyze_cohort_growth.py to silently drop all rows (dropna=True).
                "LEA Code":            "0",
                "LEA Name":            "DC Schools",
                "School Code":         school_code,
                "School Name":         school_name,
                "Assessment Name":     assess_name,
                "Subject":             subject,
                "Student Group":       student_group,
                "Student Group Value": student_group_value,
                "Tested Grade/Subject": grade_label,
                "Grade of Enrollment":  grade_label,
                "Count":               str(int(prof)) if not np.isnan(prof) else "",
                "Total Count":         str(int(total)) if not np.isnan(total) else "",
                "Percent":             str(round(pct, 4)) if not np.isnan(pct) else "",
                "Year":                year,
            }

        # PARCC/DCCAPE ELA
        records.append(make_record("ELA",  "ela_total",  "ela_prof",  "ela_pct",  assessment))
        # PARCC/DCCAPE Math
        records.append(make_record("Math", "math_total", "math_prof", "math_pct", assessment))
        # MSAA ELA (only if columns exist)
        if col_idx.get("msaa_ela_total", -1) >= 0:
            records.append(make_record("ELA",  "msaa_ela_total", "msaa_ela_prof", "msaa_ela_pct",  "MSAA"))
            records.append(make_record("Math", "msaa_math_total", "msaa_math_prof", "msaa_math_pct", "MSAA"))

    return pd.DataFrame(records) if records else pd.DataFrame()


def load_file(filepath: str, year: int, assessment: str) -> pd.DataFrame:
    """Load all demographic sheets from one workbook and return combined long-format DF."""
    print(f"\n  Loading {os.path.basename(filepath)} (year={year}, assessment={assessment}) …")
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as exc:
        print(f"    ERROR opening file: {exc}")
        return pd.DataFrame()

    sheet_dfs = []
    for sheet_name, (sg, sgv) in SHEET_SUBGROUP.items():
        df = read_wide_sheet(wb, sheet_name, year, assessment, sg, sgv)
        if not df.empty:
            sheet_dfs.append(df)
            print(f"    Sheet '{sheet_name}': {len(df):,} rows")

    wb.close()

    if not sheet_dfs:
        print("    WARNING: No sheets loaded.")
        return pd.DataFrame()

    combined = pd.concat(sheet_dfs, ignore_index=True)
    print(f"    Total rows from this file: {len(combined):,}")
    return combined


# ═════════════════════════════════════════════════════════════════════════════
# Main
# ═════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 70)
    print("DC SCHOOLS WIDE-FORMAT DATA LOADER")
    print("=" * 70)

    os.makedirs(OUTPUT_PATH, exist_ok=True)

    files = find_files()
    if not files:
        print(f"\nERROR: No matching XLSX files found under {INPUT_PATH}")
        print("Expected files containing year keywords: 2015-16, 2016-17, 2017-18, 2018-19, 2021-22, 2022-23, 2023-2024")
        sys.exit(1)

    print(f"\nFound {len(files)} file(s):")
    for fpath, year, assess in files:
        print(f"  [{year}] {assess}: {os.path.relpath(fpath, REPO_ROOT)}")

    all_dfs = []
    for fpath, year, assess in files:
        df = load_file(fpath, year, assess)
        if not df.empty:
            all_dfs.append(df)

    if not all_dfs:
        print("\nERROR: No data was loaded from any file.")
        sys.exit(1)

    print("\n" + "=" * 70)
    print("COMBINING AND CLEANING")
    print("=" * 70)

    combined = pd.concat(all_dfs, ignore_index=True)
    print(f"\nRaw combined rows: {len(combined):,}")

    # Drop rows where Percent is blank (no data at all for that school/grade/subject)
    before = len(combined)
    combined = combined[combined["Percent"] != ""].copy()
    print(f"After dropping blank-percent rows: {len(combined):,} (removed {before - len(combined):,})")

    # Ensure standard column order
    std_cols = [
        "Aggregation Level", "LEA Code", "LEA Name", "School Code", "School Name",
        "Assessment Name", "Subject", "Student Group", "Student Group Value",
        "Tested Grade/Subject", "Grade of Enrollment", "Count", "Total Count",
        "Percent", "Year",
    ]
    for col in std_cols:
        if col not in combined.columns:
            combined[col] = ""
    combined = combined[std_cols]

    print(f"\nYears present: {sorted(combined['Year'].unique())}")
    print(f"Unique schools: {combined['School Name'].nunique()}")
    print(f"Subjects: {sorted(combined['Subject'].unique())}")

    output_file = os.path.join(OUTPUT_PATH, "combined_all_years.csv")
    combined.to_csv(output_file, index=False)
    print(f"\n✓ Saved {len(combined):,} rows → {output_file}")

    # Write a simple processing report
    report_file = os.path.join(OUTPUT_PATH, "processing_report.txt")
    with open(report_file, "w") as rpt:
        rpt.write("DC Schools Wide-Format Data Processing Report\n")
        rpt.write("=" * 50 + "\n\n")
        for fpath, year, assess in files:
            rpt.write(f"✓ [{year}] {assess}: {os.path.basename(fpath)}\n")
        rpt.write(f"\nTotal rows in combined_all_years.csv: {len(combined):,}\n")
        rpt.write(f"Years: {sorted(combined['Year'].unique())}\n")
        rpt.write(f"Schools: {combined['School Name'].nunique()}\n")
    print(f"✓ Processing report → {report_file}")

    print("\n" + "=" * 70)
    print("COMPLETE — run src/analyze_cohort_growth.py next")
    print("=" * 70)


if __name__ == "__main__":
    main()
